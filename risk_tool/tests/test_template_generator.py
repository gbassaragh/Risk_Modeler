"""Tests for template generation functionality.

Tests the TemplateGenerator class and template creation for different formats.
"""

import json
import tempfile
import pytest
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from risk_tool.templates.template_generator import TemplateGenerator


class TestTemplateGenerator:
    """Test template generation functionality."""

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test files."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            yield Path(tmp_dir)

    def test_generate_transmission_line_json(self, temp_dir):
        """Test transmission line JSON template generation."""
        output_file = temp_dir / "transmission_line.json"

        TemplateGenerator.generate_template(
            "transmission_line", "json", str(output_file)
        )

        assert output_file.exists()

        # Load and validate JSON structure
        with open(output_file, "r") as f:
            template = json.load(f)

        # Validate required sections
        assert "project_info" in template
        assert "wbs_items" in template
        assert "simulation_config" in template
        assert "correlations" in template

        # Validate project info
        project_info = template["project_info"]
        assert project_info["project_type"] == "TransmissionLine"
        assert project_info["voltage_class"] == "138kV"
        assert "name" in project_info
        assert "description" in project_info

        # Validate WBS items
        wbs_items = template["wbs_items"]
        assert len(wbs_items) == 8  # Should have 8 WBS items

        # Check first WBS item structure
        wbs_item = wbs_items[0]
        required_fields = [
            "code",
            "name",
            "quantity",
            "uom",
            "unit_cost",
            "dist_unit_cost",
        ]
        for field in required_fields:
            assert field in wbs_item

        # Validate distribution structure
        dist = wbs_item["dist_unit_cost"]
        assert dist["type"] == "triangular"
        assert "low" in dist and "mode" in dist and "high" in dist

        # Validate simulation config
        sim_config = template["simulation_config"]
        assert sim_config["iterations"] == 10000
        assert sim_config["sampling_method"] == "LHS"
        assert "random_seed" in sim_config

    def test_generate_substation_json(self, temp_dir):
        """Test substation JSON template generation."""
        output_file = temp_dir / "substation.json"

        TemplateGenerator.generate_template("substation", "json", str(output_file))

        assert output_file.exists()

        # Load and validate JSON structure
        with open(output_file, "r") as f:
            template = json.load(f)

        # Validate substation-specific content
        project_info = template["project_info"]
        assert project_info["project_type"] == "Substation"
        assert project_info["voltage_class"] == "138kV/12.47kV"

        # Should have substation-specific WBS items
        wbs_items = template["wbs_items"]
        assert len(wbs_items) == 8  # Should have 8 WBS items for substation

        # Check for substation-specific items
        wbs_codes = [item["code"] for item in wbs_items]
        assert "01-SITE" in wbs_codes
        assert "06-STRUCT" in wbs_codes  # Structure code is 06-STRUCT, not 02-STRUCT
        assert "02-XFMR" in wbs_codes  # Transformer equipment

    def test_generate_transmission_line_excel(self, temp_dir):
        """Test transmission line Excel template generation."""
        output_file = temp_dir / "transmission_line.xlsx"

        TemplateGenerator.generate_template(
            "transmission_line", "excel", str(output_file)
        )

        assert output_file.exists()

        # Load and validate Excel structure
        workbook = load_workbook(output_file)

        # Check required worksheets
        expected_sheets = [
            "Project_Info",
            "WBS_Items",
            "Correlations",
            "Simulation_Config",
        ]
        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames

        # Validate Project_Info sheet
        project_sheet = workbook["Project_Info"]
        assert project_sheet["A1"].value == "Parameter"
        assert project_sheet["B1"].value == "Value"

        # Check for key project fields
        field_column = [cell.value for cell in project_sheet["A"] if cell.value]
        assert "name" in field_column
        assert "project_type" in field_column
        assert "voltage_class" in field_column

        # Validate WBS_Items sheet
        wbs_sheet = workbook["WBS_Items"]
        headers = [cell.value for cell in wbs_sheet[1]]
        # Excel headers are capitalized
        expected_headers = [
            "Code",
            "Name",
            "Quantity",
            "UoM",
            "Unit Cost",
            "Dist Type",
            "Dist Low",
            "Dist Mode",
            "Dist High",
        ]
        for header in expected_headers:
            assert header in headers

        # Should have data rows (8 WBS items + header)
        assert wbs_sheet.max_row == 9

        workbook.close()

    def test_generate_transmission_line_csv(self, temp_dir):
        """Test transmission line CSV template generation."""
        output_file = temp_dir / "transmission_line_wbs.csv"

        TemplateGenerator.generate_template(
            "transmission_line", "csv", str(output_file)
        )

        assert output_file.exists()

        # Load and validate CSV structure
        df = pd.read_csv(output_file)

        # Check required columns
        expected_columns = [
            "code",
            "name",
            "quantity",
            "uom",
            "unit_cost",
            "dist_type",
            "dist_low",
            "dist_mode",
            "dist_high",
        ]
        for col in expected_columns:
            assert col in df.columns

        # Should have 8 rows for transmission line
        assert len(df) == 8

        # Validate first row data
        first_row = df.iloc[0]
        assert first_row["code"] == "01-ROW"
        assert first_row["dist_type"] == "triangular"
        assert pd.notna(first_row["dist_low"])
        assert pd.notna(first_row["dist_mode"])
        assert pd.notna(first_row["dist_high"])

    def test_generate_substation_csv(self, temp_dir):
        """Test substation CSV template generation."""
        output_file = temp_dir / "substation_wbs.csv"

        TemplateGenerator.generate_template("substation", "csv", str(output_file))

        assert output_file.exists()

        # Load and validate CSV structure
        df = pd.read_csv(output_file)

        # Should have 8 rows for substation
        assert len(df) == 8

        # Check substation-specific codes
        codes = df["code"].tolist()
        assert "01-SITE" in codes
        assert "06-STRUCT" in codes  # Structure code is 06-STRUCT, not 02-STRUCT
        assert "02-XFMR" in codes  # Transformer equipment

    def test_invalid_category(self, temp_dir):
        """Test error handling for invalid category."""
        output_file = temp_dir / "invalid.json"

        with pytest.raises(ValueError, match="Unknown project category"):
            TemplateGenerator.generate_template(
                "invalid_category", "json", str(output_file)
            )

    def test_invalid_format(self, temp_dir):
        """Test error handling for invalid format."""
        output_file = temp_dir / "template.invalid"

        with pytest.raises(ValueError, match="Unknown format"):
            TemplateGenerator.generate_template(
                "transmission_line", "invalid_format", str(output_file)
            )

    def test_template_file_completeness(self, temp_dir):
        """Test that template files are complete and valid."""
        # Test all combinations work
        categories = ["transmission_line", "substation"]
        formats = ["json", "csv"]  # Skip excel for now since it requires openpyxl

        for category in categories:
            for format_type in formats:
                output_file = temp_dir / f"{category}_test.{format_type}"

                # Should not raise exception
                TemplateGenerator.generate_template(
                    category, format_type, str(output_file)
                )

                # File should exist and have content
                assert output_file.exists()
                assert output_file.stat().st_size > 0

    def test_json_template_completeness(self, temp_dir):
        """Test that JSON templates contain all required fields."""
        output_file = temp_dir / "complete_template.json"

        TemplateGenerator.generate_template(
            "transmission_line", "json", str(output_file)
        )

        with open(output_file, "r") as f:
            template = json.load(f)

        # Test deep structure validation
        project_info = template["project_info"]
        required_project_fields = [
            "name",
            "description",
            "project_type",
            "voltage_class",
            "length_miles",
            "region",
            "base_year",
            "currency",
            "aace_class",
        ]
        for field in required_project_fields:
            assert field in project_info, f"Missing project field: {field}"

        # Test WBS item completeness
        for i, wbs_item in enumerate(template["wbs_items"]):
            required_wbs_fields = [
                "code",
                "name",
                "quantity",
                "uom",
                "unit_cost",
                "dist_unit_cost",
                "tags",
                "indirect_factor",
            ]
            for field in required_wbs_fields:
                assert field in wbs_item, f"Missing WBS field {field} in item {i}"

        # Test correlations structure
        assert len(template["correlations"]) > 0
        corr = template["correlations"][0]
        assert "items" in corr
        assert "correlation" in corr
        assert "method" in corr

        # Test simulation config completeness
        sim_config = template["simulation_config"]
        required_sim_fields = [
            "iterations",
            "random_seed",
            "sampling_method",
            "convergence_threshold",
            "max_iterations",
            "confidence_levels",
        ]
        for field in required_sim_fields:
            assert field in sim_config, f"Missing simulation field: {field}"

    def test_excel_template_data_validation(self, temp_dir):
        """Test that Excel templates have proper data validation."""
        output_file = temp_dir / "validation_test.xlsx"

        TemplateGenerator.generate_template(
            "transmission_line", "excel", str(output_file)
        )

        workbook = load_workbook(output_file)
        wbs_sheet = workbook["WBS_Items"]

        # Check that numeric columns have numeric data
        for row in range(2, wbs_sheet.max_row + 1):
            quantity = wbs_sheet[f"C{row}"].value  # quantity column
            unit_cost = wbs_sheet[f"E{row}"].value  # unit_cost column

            assert isinstance(
                quantity, (int, float)
            ), f"Row {row} quantity is not numeric"
            assert isinstance(
                unit_cost, (int, float)
            ), f"Row {row} unit_cost is not numeric"
            assert quantity > 0, f"Row {row} quantity should be positive"
            assert unit_cost > 0, f"Row {row} unit_cost should be positive"

        workbook.close()

    def test_csv_template_data_types(self, temp_dir):
        """Test that CSV templates have proper data types."""
        output_file = temp_dir / "data_types_test.csv"

        TemplateGenerator.generate_template("substation", "csv", str(output_file))

        df = pd.read_csv(output_file)

        # Test numeric columns are properly formatted
        numeric_columns = ["quantity", "unit_cost"]
        for col in numeric_columns:
            assert pd.api.types.is_numeric_dtype(
                df[col]
            ), f"Column {col} should be numeric"
            assert (df[col] > 0).all(), f"Column {col} should have positive values"

        # Test distribution columns (may have NaN values)
        dist_columns = ["dist_low", "dist_mode", "dist_high"]
        for col in dist_columns:
            assert pd.api.types.is_numeric_dtype(
                df[col]
            ), f"Column {col} should be numeric"
            # Only check positive values where not NaN
            non_nan_values = df[col].dropna()
            if len(non_nan_values) > 0:
                assert (
                    non_nan_values > 0
                ).all(), f"Column {col} should have positive non-NaN values"

        # Test string columns
        string_columns = ["code", "name", "uom", "dist_type"]
        for col in string_columns:
            assert df[col].dtype == "object", f"Column {col} should be string type"
            assert not df[col].isna().any(), f"Column {col} should not have null values"


if __name__ == "__main__":
    pytest.main([__file__])
