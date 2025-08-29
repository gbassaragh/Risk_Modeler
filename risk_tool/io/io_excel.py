"""Excel file I/O operations for project and risk data.

Handles Excel import/export with robust column mapping and validation.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any, Optional, Tuple
import warnings
from pathlib import Path


class ExcelImporter:
    """Imports project and risk data from Excel files."""

    def __init__(self):
        """Initialize Excel importer."""
        self.default_wbs_columns = {
            "code": ["code", "wbs_code", "item_code"],
            "name": ["name", "description", "item_name"],
            "quantity": ["quantity", "qty"],
            "uom": ["uom", "unit", "units"],
            "unit_cost": ["unit_cost", "rate", "unit_rate"],
            "tags": ["tags", "categories"],
            "indirect_factor": ["indirect_factor", "indirects", "overhead_factor"],
        }

        self.default_risk_columns = {
            "id": ["id", "risk_id", "code"],
            "title": ["title", "name", "description"],
            "category": ["category", "type"],
            "probability": ["probability", "p", "likelihood"],
            "impact_mode": ["impact_mode", "mode"],
            "applies_to": ["applies_to", "wbs_codes"],
            "applies_by_tag": ["applies_by_tag", "tags"],
        }

    def import_project(self, file_path: str, sheet_name: str = "WBS") -> Dict[str, Any]:
        """Import project data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name containing WBS data

        Returns:
            Project dictionary
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            # Clean column names
            df.columns = df.columns.str.strip().str.lower()

            # Map columns
            column_mapping = self._map_wbs_columns(df.columns.tolist())

            # Process WBS items
            wbs_items = []
            for _, row in df.iterrows():
                if pd.isna(row.get(column_mapping.get("code"))):
                    continue

                wbs_item = self._process_wbs_row(row, column_mapping)
                if wbs_item:
                    wbs_items.append(wbs_item)

            # Try to read project metadata from separate sheet or use defaults
            project_metadata = self._read_project_metadata(file_path)

            project = {
                "id": project_metadata.get("id", f"PROJECT-{Path(file_path).stem}"),
                "type": project_metadata.get("type", "TransmissionLine"),
                "currency": project_metadata.get("currency", "USD"),
                "base_date": project_metadata.get("base_date", "2025-01-01"),
                "region": project_metadata.get("region", "US"),
                "wbs": wbs_items,
                "indirects_per_day": project_metadata.get("indirects_per_day", 25000.0),
            }

            if "escalation" in project_metadata:
                project["escalation"] = project_metadata["escalation"]

            return project

        except Exception as e:
            raise ValueError(f"Error importing project from {file_path}: {e}")

    def import_risks(
        self, file_path: str, sheet_name: str = "Risks"
    ) -> List[Dict[str, Any]]:
        """Import risk data from Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name containing risk data

        Returns:
            List of risk dictionaries
        """
        try:
            # Read Excel file
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            # Clean column names
            df.columns = df.columns.str.strip().str.lower()

            # Map columns
            column_mapping = self._map_risk_columns(df.columns.tolist())

            # Process risk items
            risks = []
            for _, row in df.iterrows():
                if pd.isna(row.get(column_mapping.get("id"))):
                    continue

                risk_item = self._process_risk_row(row, column_mapping)
                if risk_item:
                    risks.append(risk_item)

            return risks

        except Exception as e:
            raise ValueError(f"Error importing risks from {file_path}: {e}")

    def _map_wbs_columns(self, columns: List[str]) -> Dict[str, str]:
        """Map WBS columns to standard names."""
        mapping = {}

        for standard_name, possible_names in self.default_wbs_columns.items():
            for col in columns:
                if col in possible_names:
                    mapping[standard_name] = col
                    break

        return mapping

    def _map_risk_columns(self, columns: List[str]) -> Dict[str, str]:
        """Map risk columns to standard names."""
        mapping = {}

        for standard_name, possible_names in self.default_risk_columns.items():
            for col in columns:
                if col in possible_names:
                    mapping[standard_name] = col
                    break

        return mapping

    def _process_wbs_row(
        self, row: pd.Series, column_mapping: Dict[str, str]
    ) -> Optional[Dict[str, Any]]:
        """Process a WBS row into WBS item dictionary."""
        try:
            wbs_item = {
                "code": str(row[column_mapping["code"]]).strip(),
                "name": str(row[column_mapping["name"]]).strip(),
                "quantity": float(row[column_mapping["quantity"]]),
                "uom": str(row[column_mapping["uom"]]).strip(),
                "unit_cost": float(row[column_mapping["unit_cost"]]),
                "tags": [],
                "indirect_factor": 0.0,
            }

            # Process optional fields
            if "tags" in column_mapping and not pd.isna(row[column_mapping["tags"]]):
                tags_str = str(row[column_mapping["tags"]])
                wbs_item["tags"] = [tag.strip() for tag in tags_str.split(",")]

            if "indirect_factor" in column_mapping and not pd.isna(
                row[column_mapping["indirect_factor"]]
            ):
                wbs_item["indirect_factor"] = float(
                    row[column_mapping["indirect_factor"]]
                )

            # Process distributions if present
            dist_columns = [
                col for col in row.index if "dist_" in col or "distribution" in col
            ]
            for col in dist_columns:
                if not pd.isna(row[col]):
                    # Try to parse distribution specification
                    dist_spec = self._parse_distribution_spec(str(row[col]))
                    if dist_spec:
                        if "quantity" in col.lower():
                            wbs_item["dist_quantity"] = dist_spec
                        elif "cost" in col.lower() or "rate" in col.lower():
                            wbs_item["dist_unit_cost"] = dist_spec

            return wbs_item

        except Exception as e:
            warnings.warn(f"Error processing WBS row: {e}")
            return None

    def _process_risk_row(
        self, row: pd.Series, column_mapping: Dict[str, str]
    ) -> Optional[Dict[str, Any]]:
        """Process a risk row into risk item dictionary."""
        try:
            risk_item = {
                "id": str(row[column_mapping["id"]]).strip(),
                "title": str(row[column_mapping["title"]]).strip(),
                "category": str(row[column_mapping["category"]]).strip(),
                "probability": float(row[column_mapping["probability"]]),
                "impact_mode": str(row[column_mapping["impact_mode"]]).strip().lower(),
                "impact_dist": {
                    "type": "triangular",
                    "low": 1.0,
                    "mode": 1.1,
                    "high": 1.3,
                },  # Default
            }

            # Process applies_to
            if "applies_to" in column_mapping and not pd.isna(
                row[column_mapping["applies_to"]]
            ):
                applies_str = str(row[column_mapping["applies_to"]])
                risk_item["applies_to"] = [
                    item.strip() for item in applies_str.split(",")
                ]

            # Process applies_by_tag
            if "applies_by_tag" in column_mapping and not pd.isna(
                row[column_mapping["applies_by_tag"]]
            ):
                tags_str = str(row[column_mapping["applies_by_tag"]])
                risk_item["applies_by_tag"] = [
                    tag.strip() for tag in tags_str.split(",")
                ]

            # Process impact distribution
            impact_columns = [
                col
                for col in row.index
                if "impact" in col and ("dist" in col or "min" in col or "max" in col)
            ]
            if impact_columns:
                impact_dist = self._parse_impact_distribution(row, impact_columns)
                if impact_dist:
                    risk_item["impact_dist"] = impact_dist

            return risk_item

        except Exception as e:
            warnings.warn(f"Error processing risk row: {e}")
            return None

    def _parse_distribution_spec(self, spec: str) -> Optional[Dict[str, Any]]:
        """Parse distribution specification from string."""
        try:
            spec = spec.strip()

            if "triangular" in spec.lower():
                # Parse triangular(low, mode, high)
                import re

                match = re.search(
                    r"triangular\s*\(\s*([\d.]+)\s*,\s*([\d.]+)\s*,\s*([\d.]+)\s*\)",
                    spec,
                )
                if match:
                    return {
                        "type": "triangular",
                        "low": float(match.group(1)),
                        "mode": float(match.group(2)),
                        "high": float(match.group(3)),
                    }

            elif "normal" in spec.lower():
                # Parse normal(mean, std)
                import re

                match = re.search(r"normal\s*\(\s*([\d.]+)\s*,\s*([\d.]+)\s*\)", spec)
                if match:
                    return {
                        "type": "normal",
                        "mean": float(match.group(1)),
                        "stdev": float(match.group(2)),
                    }

            elif "pert" in spec.lower():
                # Parse pert(min, mode, max)
                import re

                match = re.search(
                    r"pert\s*\(\s*([\d.]+)\s*,\s*([\d.]+)\s*,\s*([\d.]+)\s*\)", spec
                )
                if match:
                    return {
                        "type": "pert",
                        "min": float(match.group(1)),
                        "most_likely": float(match.group(2)),
                        "max": float(match.group(3)),
                    }

            return None

        except Exception:
            return None

    def _parse_impact_distribution(
        self, row: pd.Series, impact_columns: List[str]
    ) -> Optional[Dict[str, Any]]:
        """Parse impact distribution from row data."""
        try:
            # Look for standard distribution columns
            min_col = next(
                (col for col in impact_columns if "min" in col.lower()), None
            )
            mode_col = next(
                (
                    col
                    for col in impact_columns
                    if ("mode" in col.lower() or "most_likely" in col.lower())
                ),
                None,
            )
            max_col = next(
                (col for col in impact_columns if "max" in col.lower()), None
            )

            if min_col and mode_col and max_col:
                return {
                    "type": "pert",
                    "min": float(row[min_col]),
                    "most_likely": float(row[mode_col]),
                    "max": float(row[max_col]),
                }

            # Look for triangular distribution
            low_col = next(
                (col for col in impact_columns if "low" in col.lower()), None
            )
            high_col = next(
                (col for col in impact_columns if "high" in col.lower()), None
            )

            if low_col and mode_col and high_col:
                return {
                    "type": "triangular",
                    "low": float(row[low_col]),
                    "mode": float(row[mode_col]),
                    "high": float(row[high_col]),
                }

            return None

        except Exception:
            return None

    def _read_project_metadata(self, file_path: str) -> Dict[str, Any]:
        """Read project metadata from Excel file."""
        metadata = {}

        try:
            # Try to read from 'Project' sheet
            project_df = pd.read_excel(file_path, sheet_name="Project")

            if len(project_df.columns) >= 2:
                # Assume first column is keys, second is values
                for _, row in project_df.iterrows():
                    key = str(row.iloc[0]).strip().lower()
                    value = row.iloc[1]

                    if not pd.isna(value):
                        metadata[key] = value

        except Exception:
            # Use defaults if Project sheet not found
            pass

        return metadata


class ExcelExporter:
    """Exports results and templates to Excel files."""

    def export_results(self, results_dict: Dict[str, Any], file_path: str) -> None:
        """Export simulation results to Excel file.

        Args:
            results_dict: Dictionary containing all results
            file_path: Output file path
        """
        try:
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Summary sheet
                if "summary" in results_dict:
                    summary_df = self._create_summary_dataframe(results_dict["summary"])
                    summary_df.to_excel(writer, sheet_name="Summary", index=False)

                # WBS breakdown
                if "wbs_statistics" in results_dict:
                    wbs_df = pd.DataFrame(results_dict["wbs_statistics"]).T
                    wbs_df.to_excel(writer, sheet_name="WBS_Breakdown")

                # Risk contributions
                if "risk_contributions" in results_dict:
                    risk_df = pd.DataFrame(results_dict["risk_contributions"]).T
                    risk_df.to_excel(writer, sheet_name="Risk_Analysis")

                # Raw simulation data (limited)
                if "simulation_data" in results_dict:
                    sim_data = results_dict["simulation_data"]
                    if isinstance(sim_data, dict) and "total_costs" in sim_data:
                        # Export sample of simulation data
                        sample_size = min(10000, len(sim_data["total_costs"]))
                        indices = np.linspace(
                            0, len(sim_data["total_costs"]) - 1, sample_size, dtype=int
                        )

                        sample_df = pd.DataFrame(
                            {"total_cost": sim_data["total_costs"][indices]}
                        )

                        sample_df.to_excel(
                            writer, sheet_name="Simulation_Sample", index=False
                        )

                # Format sheets
                self._format_excel_sheets(writer)

        except Exception as e:
            raise ValueError(f"Error exporting results to {file_path}: {e}")

    def create_wbs_template(self, file_path: str) -> None:
        """Create WBS template Excel file.

        Args:
            file_path: Output file path
        """
        try:
            wb = Workbook()

            # WBS sheet
            ws = wb.active
            ws.title = "WBS"

            # Headers
            headers = [
                "code",
                "name",
                "quantity",
                "uom",
                "unit_cost",
                "tags",
                "indirect_factor",
                "dist_quantity",
                "dist_unit_cost",
            ]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Example data
            example_data = [
                [
                    "1.1",
                    "ROW Clearing & Access",
                    10.0,
                    "miles",
                    120000.0,
                    "TL,access,weather_sensitive",
                    0.12,
                    "triangular(9,10,12)",
                    "pert(100000,120000,160000)",
                ],
                [
                    "2.1",
                    "Steel Structures",
                    50,
                    "each",
                    25000.0,
                    "structures,steel",
                    0.08,
                    "",
                    "lognormal(25000,0.2)",
                ],
                [
                    "2.2",
                    "Foundations",
                    50,
                    "each",
                    15000.0,
                    "foundations,concrete",
                    0.10,
                    "",
                    "triangular(12000,15000,20000)",
                ],
                [
                    "2.3",
                    "Conductor & OPGW",
                    60.0,
                    "miles-conductor",
                    85000.0,
                    "materials,commodity_aluminum",
                    0.05,
                    "",
                    "lognormal(85000,0.15)",
                ],
            ]

            for row, data in enumerate(example_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)

            # Project metadata sheet
            proj_ws = wb.create_sheet("Project")
            proj_ws.append(["Property", "Value"])
            proj_ws.append(["id", "TL-115kV-10mi-2025-CT"])
            proj_ws.append(["type", "TransmissionLine"])
            proj_ws.append(["currency", "USD"])
            proj_ws.append(["base_date", "2025-06-01"])
            proj_ws.append(["region", "ISO-NE"])
            proj_ws.append(["indirects_per_day", 25000.0])

            # Format
            self._format_template_sheets(wb)

            wb.save(file_path)

        except Exception as e:
            raise ValueError(f"Error creating WBS template at {file_path}: {e}")

    def create_risk_template(self, file_path: str) -> None:
        """Create risk template Excel file.

        Args:
            file_path: Output file path
        """
        try:
            wb = Workbook()

            # Risks sheet
            ws = wb.active
            ws.title = "Risks"

            # Headers
            headers = [
                "id",
                "title",
                "category",
                "probability",
                "impact_mode",
                "impact_min",
                "impact_most_likely",
                "impact_max",
                "applies_to",
                "applies_by_tag",
                "schedule_days_min",
                "schedule_days_max",
            ]

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Example data with T&D specific risks
            example_data = [
                [
                    "R-ENV-001",
                    "Wetland matting extent > plan",
                    "Environmental/Access",
                    0.45,
                    "multiplicative",
                    1.03,
                    1.12,
                    1.35,
                    "1.1",
                    "",
                    0,
                    20,
                ],
                [
                    "R-PROC-002",
                    "Steel pole procurement volatility",
                    "Commodity/Market",
                    0.35,
                    "multiplicative",
                    0.9,
                    1.0,
                    1.3,
                    "",
                    "steel,structures",
                    0,
                    0,
                ],
                [
                    "R-OUT-003",
                    "Outage window constraints",
                    "Operations",
                    0.25,
                    "additive",
                    200000,
                    400000,
                    1200000,
                    "",
                    "indirects",
                    0,
                    30,
                ],
                [
                    "R-WX-004",
                    "Storm season productivity loss",
                    "Weather",
                    0.60,
                    "multiplicative",
                    1.1,
                    1.25,
                    1.5,
                    "",
                    "weather_sensitive",
                    5,
                    15,
                ],
            ]

            for row, data in enumerate(example_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)

            # Correlations sheet
            corr_ws = wb.create_sheet("Correlations")
            corr_ws.append(["Variable1", "Variable2", "Correlation", "Method"])
            corr_ws.append(["commodity_aluminum", "Conductor & OPGW", 0.6, "spearman"])
            corr_ws.append(["weather_sensitive", "indirects_per_day", 0.4, "spearman"])

            # Format
            self._format_template_sheets(wb)

            wb.save(file_path)

        except Exception as e:
            raise ValueError(f"Error creating risk template at {file_path}: {e}")

    def _create_summary_dataframe(self, summary: Dict[str, Any]) -> pd.DataFrame:
        """Create summary DataFrame from results."""
        data = []

        for key, value in summary.items():
            if isinstance(value, (int, float)):
                data.append({"Metric": key, "Value": value})
            elif isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    if isinstance(sub_value, (int, float)):
                        data.append({"Metric": f"{key}_{sub_key}", "Value": sub_value})

        return pd.DataFrame(data)

    def _format_excel_sheets(self, writer):
        """Format Excel sheets with styling."""
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(50, max(12, max_length + 2))
                ws.column_dimensions[column_letter].width = adjusted_width

    def _format_template_sheets(self, wb):
        """Format template sheets with styling."""
        for ws in wb.worksheets:
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(30, max(12, max_length + 2))
                ws.column_dimensions[column_letter].width = adjusted_width
