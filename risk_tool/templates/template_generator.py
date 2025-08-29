"""Template generation utilities for Excel, CSV, and JSON formats.

Creates professional project templates for transmission lines and substations
with proper formatting, data validation, and user guidance.
"""

from pathlib import Path
from typing import Dict, List, Any, Optional
import json
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None

from ..core.logging_config import get_logger

logger = get_logger(__name__)


class TemplateGenerator:
    """Generate project templates in various formats."""

    def __init__(self):
        """Initialize template generator."""
        self.templates_dir = Path(__file__).parent

    def create_transmission_line_excel(self, output_path: Path) -> None:
        """Create Excel template for transmission line projects."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available - cannot create Excel templates")

        # Load JSON template
        with open(self.templates_dir / "transmission_line_template.json") as f:
            template_data = json.load(f)

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet and create custom sheets
        wb.remove(wb.active)

        # Create sheets
        self._create_project_info_sheet(wb, template_data["project_info"])
        self._create_wbs_sheet(wb, template_data["wbs_items"], "Transmission Line")
        self._create_correlations_sheet(wb, template_data["correlations"])
        self._create_config_sheet(wb, template_data["simulation_config"])
        self._create_instructions_sheet(wb, "transmission_line")

        # Save workbook
        wb.save(output_path)
        logger.info(f"Created transmission line Excel template: {output_path}")

    def create_substation_excel(self, output_path: Path) -> None:
        """Create Excel template for substation projects."""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available - cannot create Excel templates")

        # Load JSON template
        with open(self.templates_dir / "substation_template.json") as f:
            template_data = json.load(f)

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create sheets
        self._create_project_info_sheet(wb, template_data["project_info"])
        self._create_wbs_sheet(wb, template_data["wbs_items"], "Substation")
        self._create_correlations_sheet(wb, template_data["correlations"])
        self._create_config_sheet(wb, template_data["simulation_config"])
        self._create_instructions_sheet(wb, "substation")

        wb.save(output_path)
        logger.info(f"Created substation Excel template: {output_path}")

    def _create_project_info_sheet(
        self, wb: openpyxl.Workbook, project_info: Dict[str, Any]
    ) -> None:
        """Create project information sheet."""
        ws = wb.create_sheet("Project_Info")

        # Headers
        ws["A1"] = "Parameter"
        ws["B1"] = "Value"
        ws["C1"] = "Description"

        # Apply header formatting
        for cell in ws["A1:C1"][0]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Add project info
        row = 2
        descriptions = {
            "name": "Project name or identifier",
            "description": "Brief project description",
            "project_type": "TransmissionLine or Substation",
            "voltage_class": "Operating voltage (e.g., 138kV, 138kV/12.47kV)",
            "length_miles": "Line length in miles (transmission only)",
            "capacity_mva": "Transformer capacity in MVA (substation only)",
            "circuit_count": "Number of circuits (transmission only)",
            "bay_count": "Number of bays (substation only)",
            "terrain_type": "flat, hilly, or mixed",
            "substation_type": "transmission or distribution",
            "region": "Geographic region",
            "base_year": "Base cost year",
            "currency": "Cost currency (USD, CAD, EUR)",
            "aace_class": "AACE estimate class (1-5)",
            "contingency_target": "Target contingency percentage",
        }

        for key, value in project_info.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            ws[f"C{row}"] = descriptions.get(key, "")
            row += 1

        # Auto-fit columns
        self._autofit_columns(ws)

    def _create_wbs_sheet(
        self, wb: openpyxl.Workbook, wbs_items: List[Dict[str, Any]], project_type: str
    ) -> None:
        """Create WBS items sheet."""
        ws = wb.create_sheet("WBS_Items")

        # Headers
        headers = [
            "Code",
            "Name",
            "Quantity",
            "UoM",
            "Unit Cost",
            "Dist Type",
            "Dist Low",
            "Dist Mode",
            "Dist High",
            "Dist Min",
            "Dist Most Likely",
            "Dist Max",
            "Dist Mean",
            "Dist StdDev",
            "Truncate Low",
            "Truncate High",
            "Tags",
            "Indirect Factor",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Add WBS items
        for row, item in enumerate(wbs_items, 2):
            ws[f"A{row}"] = item["code"]
            ws[f"B{row}"] = item["name"]
            ws[f"C{row}"] = item["quantity"]
            ws[f"D{row}"] = item["uom"]
            ws[f"E{row}"] = item["unit_cost"]

            # Distribution parameters
            if "dist_unit_cost" in item and item["dist_unit_cost"]:
                dist = item["dist_unit_cost"]
                ws[f"F{row}"] = dist["type"]

                if dist["type"] == "triangular":
                    ws[f"G{row}"] = dist.get("low", "")
                    ws[f"H{row}"] = dist.get("mode", "")
                    ws[f"I{row}"] = dist.get("high", "")
                elif dist["type"] == "pert":
                    ws[f"J{row}"] = dist.get("min", "")
                    ws[f"K{row}"] = dist.get("most_likely", "")
                    ws[f"L{row}"] = dist.get("max", "")
                elif dist["type"] == "normal":
                    ws[f"M{row}"] = dist.get("mean", "")
                    ws[f"N{row}"] = dist.get("stdev", "")
                    ws[f"O{row}"] = dist.get("truncate_low", "")
                    ws[f"P{row}"] = dist.get("truncate_high", "")

            ws[f"Q{row}"] = ",".join(item.get("tags", []))
            ws[f"R{row}"] = item.get("indirect_factor", 0.0)

        # Add data validation for distribution types
        dist_validation = DataValidation(
            type="list",
            formula1='"triangular,pert,normal,lognormal,uniform"',
            allow_blank=True,
        )
        ws.add_data_validation(dist_validation)
        dist_validation.add(f"F2:F{len(wbs_items) + 1}")

        self._autofit_columns(ws)

    def _create_correlations_sheet(
        self, wb: openpyxl.Workbook, correlations: List[Dict[str, Any]]
    ) -> None:
        """Create correlations sheet."""
        ws = wb.create_sheet("Correlations")

        # Headers
        headers = ["Item 1", "Item 2", "Correlation", "Method", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Add correlations
        for row, corr in enumerate(correlations, 2):
            ws[f"A{row}"] = corr["items"][0]
            ws[f"B{row}"] = corr["items"][1]
            ws[f"C{row}"] = corr["correlation"]
            ws[f"D{row}"] = corr.get("method", "spearman")
            ws[f"E{row}"] = (
                f"Correlation between {corr['items'][0]} and {corr['items'][1]}"
            )

        # Add validation for correlation values (-1 to 1)
        corr_validation = DataValidation(
            type="decimal", operator="between", formula1=-1, formula2=1
        )
        ws.add_data_validation(corr_validation)
        corr_validation.add(f"C2:C{len(correlations) + 10}")

        # Add validation for methods
        method_validation = DataValidation(type="list", formula1='"spearman,pearson"')
        ws.add_data_validation(method_validation)
        method_validation.add(f"D2:D{len(correlations) + 10}")

        self._autofit_columns(ws)

    def _create_config_sheet(
        self, wb: openpyxl.Workbook, config: Dict[str, Any]
    ) -> None:
        """Create simulation configuration sheet."""
        ws = wb.create_sheet("Simulation_Config")

        # Headers
        ws["A1"] = "Parameter"
        ws["B1"] = "Value"
        ws["C1"] = "Description"

        for cell in ws["A1:C1"][0]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Configuration parameters
        descriptions = {
            "iterations": "Number of Monte Carlo iterations",
            "random_seed": "Random seed for reproducibility",
            "sampling_method": "Sampling method (LHS or MCS)",
            "convergence_threshold": "Convergence threshold for stopping",
            "max_iterations": "Maximum iterations allowed",
            "confidence_levels": "Confidence levels for percentiles",
        }

        row = 2
        for key, value in config.items():
            ws[f"A{row}"] = key
            if key == "confidence_levels":
                ws[f"B{row}"] = ",".join(map(str, value))
            else:
                ws[f"B{row}"] = value
            ws[f"C{row}"] = descriptions.get(key, "")
            row += 1

        self._autofit_columns(ws)

    def _create_instructions_sheet(
        self, wb: openpyxl.Workbook, project_type: str
    ) -> None:
        """Create instructions sheet."""
        ws = wb.create_sheet("Instructions")

        # Title
        ws["A1"] = f"{project_type.title()} Project Template - Instructions"
        ws["A1"].font = Font(size=16, bold=True, color="366092")

        # Instructions content
        instructions = [
            "",
            "OVERVIEW:",
            "This template helps you set up Monte Carlo risk analysis for utility T&D projects.",
            "Complete each sheet with your project-specific data.",
            "",
            "SHEET DESCRIPTIONS:",
            "",
            "1. Project_Info:",
            "   - Enter basic project information",
            "   - All fields are required for proper analysis",
            "",
            "2. WBS_Items:",
            "   - Work Breakdown Structure with cost estimates",
            "   - Enter base costs and uncertainty distributions",
            "   - Use distribution types: triangular, pert, normal, lognormal, uniform",
            "   - Tags help categorize items for analysis",
            "",
            "3. Correlations:",
            "   - Define correlations between WBS items",
            "   - Values must be between -1 and 1",
            "   - Use spearman for rank correlations (recommended)",
            "",
            "4. Simulation_Config:",
            "   - Monte Carlo simulation parameters",
            "   - 10,000+ iterations recommended for stable results",
            "   - LHS sampling provides better convergence than MCS",
            "",
            "DISTRIBUTION TYPES:",
            "",
            "• Triangular: low, mode, high",
            "  Good for expert estimates with min/most-likely/max",
            "",
            "• PERT: min, most_likely, max",
            "  Beta distribution with emphasis on most likely",
            "",
            "• Normal: mean, stdev, truncate_low, truncate_high",
            "  Symmetric uncertainty, often truncated for costs",
            "",
            "• Lognormal: mean, sigma",
            "  For multiplicative processes, right-skewed",
            "",
            "• Uniform: low, high",
            "  Equal probability across range",
            "",
            "NEXT STEPS:",
            "",
            "1. Complete all sheets with your project data",
            "2. Save the file with a descriptive name",
            "3. Run: risk-tool run your_file.xlsx",
            "4. Review results in the output directory",
            "",
            "For questions, see USER_GUIDE.md or run: risk-tool --help",
        ]

        for row, instruction in enumerate(instructions, 2):
            ws[f"A{row}"] = instruction
            if instruction.endswith(":"):
                ws[f"A{row}"].font = Font(bold=True)

        self._autofit_columns(ws)

    def _autofit_columns(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def generate_template(category: str, format_type: str, output_path: str) -> None:
        """Generate template in specified format.

        Args:
            category: Project category ('transmission_line' or 'substation')
            format_type: Output format ('json', 'excel', 'csv')
            output_path: Path for output file
        """
        generator = TemplateGenerator()
        output_file = Path(output_path)

        # Create output directory if it doesn't exist
        output_file.parent.mkdir(parents=True, exist_ok=True)

        if category == "transmission_line":
            if format_type == "json":
                generator._create_json_template("transmission_line", output_file)
            elif format_type == "excel":
                generator.create_transmission_line_excel(output_file)
            elif format_type == "csv":
                generator._create_csv_template("transmission_line", output_file)
            else:
                raise ValueError(f"Unknown format: {format_type}")

        elif category == "substation":
            if format_type == "json":
                generator._create_json_template("substation", output_file)
            elif format_type == "excel":
                generator.create_substation_excel(output_file)
            elif format_type == "csv":
                generator._create_csv_template("substation", output_file)
            else:
                raise ValueError(f"Unknown format: {format_type}")
        else:
            raise ValueError(f"Unknown project category: {category}")

        logger.info(f"Created {category} {format_type} template: {output_path}")

    def _create_json_template(self, category: str, output_path: Path) -> None:
        """Create JSON template by copying existing template file."""
        template_file = self.templates_dir / f"{category}_template.json"
        if not template_file.exists():
            raise FileNotFoundError(f"Template file not found: {template_file}")

        # Copy template to output location
        import shutil

        shutil.copy2(template_file, output_path)

    def _create_csv_template(self, category: str, output_path: Path) -> None:
        """Create CSV template by copying existing CSV file."""
        csv_file = self.templates_dir / f"{category}_wbs.csv"
        if not csv_file.exists():
            raise FileNotFoundError(f"CSV template file not found: {csv_file}")

        # Copy CSV template to output location
        import shutil

        shutil.copy2(csv_file, output_path)


def create_all_templates(output_dir: Path) -> None:
    """Create all template files in the specified directory."""
    generator = TemplateGenerator()

    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        # Create Excel templates
        if EXCEL_AVAILABLE:
            generator.create_transmission_line_excel(
                output_dir / "transmission_line_template.xlsx"
            )
            generator.create_substation_excel(output_dir / "substation_template.xlsx")
        else:
            logger.warning("openpyxl not available - skipping Excel templates")

        logger.info(f"All templates created in {output_dir}")

    except Exception as e:
        logger.error(f"Error creating templates: {e}")
        raise


if __name__ == "__main__":
    # Create templates in current directory
    create_all_templates(Path.cwd() / "generated_templates")
